# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import font
from dataclasses import dataclass
import openpyxl
import random


@dataclass
class config:
    # エクセルファイルのパスとシート名を指定 
    file_name =   'FE単語意味頑張るマン.xlsx'
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0] 
    titlefont=("メイリオ","64")
    checkbtn_font=("メイリオ","20")
    message_font= ("メイリオ","12")
    default_button_width = 20
    default_button_padding = 8
    chapter_1_text = "第一章　コンピュータの構成要素"
    chapter_2_text = "第二章　ソフトウェアとマルチメディア"
    chapter_3_text = "第三章　基礎理論"
    chapter_4_text = "第四章　アルゴリズムとプログラミング"
    chapter_5_text = "第五章　システム構成要素"
    chapter_6_text = "第六章　データベース技術"
    chapter_7_text = "第七章　ネットワーク技術"
    chapter_8_text = "第八章　情報セキュリティ"
    chapter_9_text = "第九章　システム開発技術"
    chapter_10_text = "第十章　マネジメント系"
    chapter_11_text = "第十一章　ストラテジ系"
    chapter_list = [    "第一章　コンピュータの構成要素",    "第二章　ソフトウェアとマルチメディア",    "第三章　基礎理論",    "第四章　アルゴリズムとプログラミング",    "第五章　システム構成要素",    "第六章　データベース技術",    "第七章　ネットワーク技術",    "第八章　情報セキュリティ",    "第九章　システム開発技術",    "第十章　マネジメント系",    "第十一章　ストラテジ系"]

class Application_index(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=1344, height=756, borderwidth=1, relief='solid')
        self.root = root
        self.pack(anchor='center',expand=1)
        self.pack_propagate(0)
        self.create_widgets()
        root.bind("q", lambda event: self.change_quest_setting() if self.quest_setting_btn.winfo_ismapped() else None)
        root.bind("w", lambda event: self.change_grades() if self.grades_btn.winfo_ismapped() else None)
        root.bind("e", lambda event: self.change_data() if self.data_btn.winfo_ismapped() else None)
        root.bind("c", lambda event: self.root.destroy() if self.return_btn.winfo_ismapped() else None)

    def create_widgets(self):     
        style = ttk.Style()
        style.configure("default.TButton", font=("メイリオ", 16))
        
        # タイトルラベル
        label = tk.Label(self)
        label['text'] = 'FE単語意味頑張るマン'
        label['font'] = config.titlefont
        label.pack()

        # 閉じるボタン
        self.quit_btn = ttk.Button(self)
        self.quit_btn['text'] = '終了する'
        self.quit_btn['style'] = "default.TButton"
        self.quit_btn['width'] = config.default_button_width
        self.quit_btn['padding'] = config.default_button_padding
        self.quit_btn['command'] = self.root.destroy
        self.quit_btn.pack(side='bottom',anchor='sw')        

        # 問題を解くボタン
        self.quest_setting_btn = ttk.Button(self)
        self.quest_setting_btn['text'] = '問題を解く(q)'
        self.quest_setting_btn['style'] = "default.TButton"
        self.quest_setting_btn['width'] = config.default_button_width
        self.quest_setting_btn['padding'] = config.default_button_padding
        self.quest_setting_btn['command'] = self.change_quest_setting
        self.quest_setting_btn.bind('<Return>', self.change_quest_setting)
        self.quest_setting_btn.pack()
        
        # 成績を見るボタン
        self.grades_btn = ttk.Button(self)
        self.grades_btn['text'] = '成績を見る(w)'
        self.grades_btn['style'] = "default.TButton"
        self.grades_btn['width'] = config.default_button_width
        self.grades_btn['padding'] = config.default_button_padding
        self.grades_btn['command'] = self.change_grades
        self.grades_btn.bind('<Return>', self.change_grades)
        self.grades_btn.pack()

        # データを変更するボタン
        self.data_btn = ttk.Button(self)
        self.data_btn['text'] = '成績を見る(e)'
        self.data_btn['style'] = "default.TButton"
        self.data_btn['width'] = config.default_button_width
        self.data_btn['padding'] = config.default_button_padding
        self.data_btn['command'] = self.change_data
        self.data_btn.bind('<Return>', self.change_data)
        self.data_btn.pack()

    def change_quest_setting(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_quest_setting(self.root)
        app.pack()

    def change_grades(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_grades(self.root)
        app.pack()

    def change_data(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_data(self.root)
        app.pack()

class Application_quest_setting(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=1344, height=756, borderwidth=1, relief='solid')
        self.root = root
        self.pack(anchor='center',expand=1)
        self.pack_propagate(0)
        self.create_widgets()
        self.quest_mode = 0
        root.bind("m", lambda event: self.change_index() if self.return_btn.winfo_ismapped() else None)
        root.bind("q", lambda event: self.yougo_to_imi_btn_click() if self.yougo_to_imi_btn.winfo_ismapped() else None)
        root.bind("w", lambda event: self.imi_to_yougo_btn_click() if self.imi_to_yougo_btn.winfo_ismapped() else None)
        root.bind("e", lambda event: self.mix_btn_click() if self.mix_btn.winfo_ismapped() else None)

    def create_widgets(self):     
        style = ttk.Style()
        style.configure("default.TButton", font=("メイリオ", 16))
        
        # タイトルラベル
        self.warning_label = tk.Label(self)
        self.warning_label['text'] = ''
        self.warning_label['font'] = config.checkbtn_font
        self.warning_label.pack()

        # メニューに戻るボタン
        self.return_btn = ttk.Button(self)
        self.return_btn['text'] = 'メニューに戻る(m)'
        self.return_btn['style'] = "default.TButton"
        self.return_btn['width'] = config.default_button_width
        self.return_btn['padding'] = config.default_button_padding
        self.return_btn['command'] = self.change_index
        self.return_btn.bind('<Return>', self.change_index)
        self.return_btn.pack(side='bottom',anchor='sw')        

        # 用語to意味ボタン
        self.yougo_to_imi_btn = ttk.Button(self)
        self.yougo_to_imi_btn['text'] = '用語から説明を答える(q)'
        self.yougo_to_imi_btn['style'] = "default.TButton"
        self.yougo_to_imi_btn['width'] = config.default_button_width
        self.yougo_to_imi_btn['padding'] = config.default_button_padding
        self.yougo_to_imi_btn['command'] = self.yougo_to_imi_btn_click
        self.yougo_to_imi_btn.bind('<Return>', self.yougo_to_imi_btn_click)
        self.yougo_to_imi_btn.place(x='1058', y='40')
        

        # 意味to用語ボタン
        self.imi_to_yougo_btn = ttk.Button(self)
        self.imi_to_yougo_btn['text'] = '説明から用語を答える(w)'
        self.imi_to_yougo_btn['style'] = "default.TButton"
        self.imi_to_yougo_btn['width'] = config.default_button_width
        self.imi_to_yougo_btn['padding'] = config.default_button_padding
        self.imi_to_yougo_btn['command'] = self.imi_to_yougo_btn_click
        self.imi_to_yougo_btn.bind('<Return>', self.imi_to_yougo_btn_click)
        self.imi_to_yougo_btn.place(x='1058', y='120')

        # ミックスボタン
        self.mix_btn = ttk.Button(self)
        self.mix_btn['text'] = 'ミックス！(e)'
        self.mix_btn['style'] = "default.TButton"
        self.mix_btn['width'] = config.default_button_width
        self.mix_btn['padding'] = config.default_button_padding
        self.mix_btn['command'] = self.mix_btn_click
        self.mix_btn.bind('<Return>', self.mix_btn_click)
        self.mix_btn.place(x='1058', y='200')

        # チェックボックスすべてのブーリアンバー
        self.is_checklist =[tk.BooleanVar() for _ in range(11)] 

        # チェックボタン1
        check_btn_1 = tk.Checkbutton(self)
        check_btn_1['text'] = config.chapter_1_text
        check_btn_1['font'] = config.checkbtn_font
        check_btn_1['variable'] = self.is_checklist[0]
        check_btn_1.place(x='40', y='40')

        # チェックボタン2
        check_btn_2 = tk.Checkbutton(self)
        check_btn_2['text'] = config.chapter_2_text
        check_btn_2['font'] = config.checkbtn_font
        check_btn_2['variable'] = self.is_checklist[1]
        check_btn_2.place(x='40', y='90')
        
        # チェックボタン3
        check_btn_3 = tk.Checkbutton(self)
        check_btn_3['text'] = config.chapter_3_text
        check_btn_3['font'] = config.checkbtn_font
        check_btn_3['variable'] = self.is_checklist[2]
        check_btn_3.place(x='40', y='140')

        # チェックボタン4
        check_btn_4 = tk.Checkbutton(self)
        check_btn_4['text'] = config.chapter_4_text
        check_btn_4['font'] = config.checkbtn_font
        check_btn_4['variable'] = self.is_checklist[3]
        check_btn_4.place(x='40', y='190')

        # チェックボタン5
        check_btn_5 = tk.Checkbutton(self)
        check_btn_5['text'] = config.chapter_5_text
        check_btn_5['font'] = config.checkbtn_font
        check_btn_5['variable'] = self.is_checklist[4]
        check_btn_5.place(x='40', y='240')

        # チェックボタン6
        check_btn_6 = tk.Checkbutton(self)
        check_btn_6['text'] = config.chapter_6_text
        check_btn_6['font'] = config.checkbtn_font
        check_btn_6['variable'] = self.is_checklist[5]
        check_btn_6.place(x='40', y='290')
        
        # チェックボタン7
        check_btn_7 = tk.Checkbutton(self)
        check_btn_7['text'] = config.chapter_7_text
        check_btn_7['font'] = config.checkbtn_font
        check_btn_7['variable'] = self.is_checklist[6]
        check_btn_7.place(x='40', y='340')

        # チェックボタン8
        check_btn_8 = tk.Checkbutton(self)
        check_btn_8['text'] = config.chapter_8_text
        check_btn_8['font'] = config.checkbtn_font
        check_btn_8['variable'] = self.is_checklist[7]
        check_btn_8.place(x='40', y='390')

        # チェックボタン9
        check_btn_9 = tk.Checkbutton(self)
        check_btn_9['text'] = config.chapter_9_text
        check_btn_9['font'] = config.checkbtn_font
        check_btn_9['variable'] = self.is_checklist[8]
        check_btn_9.place(x='40', y='440')
        
        # チェックボタン10
        check_btn_10 = tk.Checkbutton(self)
        check_btn_10['text'] = config.chapter_10_text
        check_btn_10['font'] = config.checkbtn_font
        check_btn_10['variable'] = self.is_checklist[9]
        check_btn_10.place(x='40', y='490')

        # チェックボタン11
        check_btn_11 = tk.Checkbutton(self)
        check_btn_11['text'] = config.chapter_11_text
        check_btn_11['font'] = config.checkbtn_font
        check_btn_11['variable'] = self.is_checklist[10]
        check_btn_11.place(x='40', y='540')

    # どれにもチェックを入れてないの防止する
    def check_checcker(self):
        for i in self.is_checklist:
            if i.get():
                return True
        return False  

    def change_index(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_index(self.root)
        app.pack()

    def change_quest(self,event=None):
        if not self.check_checcker():
            self.warning_label['text'] = 'どれかにチェックを入れて～(´；ω；`)ｳｩｩ'
        else:
            
            # 現在のフレームを非表示にする
            self.pack_forget()
            # 新しいフレームを作成して表示する
            app = Application_quest(self.root ,is_checklist = self.is_checklist ,quest_mode = self.quest_mode)
            app.pack()

    # ボタンを押したときの挙動
    def yougo_to_imi_btn_click(self,event=None):
        self.quest_mode = 1
        self.change_quest()

    def imi_to_yougo_btn_click(self,event=None):
        self.quest_mode = 2
        self.change_quest()

    def mix_btn_click(self,event=None):
        self.quest_mode = 3
        self.change_quest()

class Application_quest(tk.Frame):
    def __init__(self, root=None, is_checklist = None ,quest_mode = None):
        super().__init__(root, width=1344, height=756, borderwidth=1, relief='solid')
        self.root = root
        self.pack(anchor='center',expand=1)
        self.pack_propagate(0)
        self.create_widgets()
        root.bind("q", lambda event: self.create_quest() if self.next_btn.winfo_ismapped() else None)
        root.bind("m", lambda event: self.change_index() if self.return_btn.winfo_ismapped() else None)
        root.bind("1", lambda event: self.choose_answer(num=0) if self.answer_1_btn.winfo_ismapped() else None)
        root.bind("2", lambda event: self.choose_answer(num=1) if self.answer_2_btn.winfo_ismapped() else None)
        root.bind("3", lambda event: self.choose_answer(num=2) if self.answer_3_btn.winfo_ismapped() else None)
        root.bind("4", lambda event: self.choose_answer(num=3) if self.answer_4_btn.winfo_ismapped() else None)
        self.file_name = 'FE単語意味頑張るマン.xlsx'
        self.is_checklist = is_checklist
        self.decision_answers = []
        self.decision_quest = []
        self.quest_mode = quest_mode
        print("クエストモードは" + str(self.quest_mode))
        # 初期問題作成
        self.create_quest()

    def create_widgets(self):     
        style = ttk.Style()
        style.configure("default.TButton", font=("メイリオ", 16))
        
        # タイトルラベル
        self.label = tk.Label(self)
        self.label['text'] = ''
        self.label['font'] = config.checkbtn_font
        self.label.pack()

        # メニューに戻るボタン
        self.return_btn = ttk.Button(self)
        self.return_btn['text'] = 'メニューに戻る(m)'
        self.return_btn['style'] = "default.TButton"
        self.return_btn['width'] = config.default_button_width
        self.return_btn['padding'] = config.default_button_padding
        self.return_btn['command'] = self.change_index
        self.return_btn.bind('<Return>', self.change_index)
        self.return_btn.pack(side='bottom',anchor='sw')              

        # １ボタン
        self.answer_1_btn = ttk.Button(self)
        self.answer_1_btn['text'] = '1'
        self.answer_1_btn['style'] = "default.TButton"
        self.answer_1_btn['width'] = config.default_button_width
        self.answer_1_btn['padding'] = config.default_button_padding
        self.answer_1_btn['command'] = lambda: self.choose_answer(num=0)
        self.answer_1_btn.bind('<Return>', lambda: self.choose_answer(num=0))

        # 2ボタン
        self.answer_2_btn = ttk.Button(self)
        self.answer_2_btn['text'] = '2'
        self.answer_2_btn['style'] = "default.TButton"
        self.answer_2_btn['width'] = config.default_button_width
        self.answer_2_btn['padding'] = config.default_button_padding
        self.answer_2_btn['command'] = lambda: self.choose_answer(num=1)
        self.answer_2_btn.bind('<Return>', lambda: self.choose_answer(num=1))

        # 3ボタン
        self.answer_3_btn = ttk.Button(self)
        self.answer_3_btn['text'] = '3'
        self.answer_3_btn['style'] = "default.TButton"
        self.answer_3_btn['width'] = config.default_button_width
        self.answer_3_btn['padding'] = config.default_button_padding
        self.answer_3_btn['command'] = lambda: self.choose_answer(num=2)
        self.answer_3_btn.bind('<Return>', lambda: self.choose_answer(num=2))

        # 4ボタン
        self.answer_4_btn = ttk.Button(self)
        self.answer_4_btn['text'] = '4'
        self.answer_4_btn['style'] = "default.TButton"
        self.answer_4_btn['width'] = config.default_button_width
        self.answer_4_btn['padding'] = config.default_button_padding
        self.answer_4_btn['command'] = lambda: self.choose_answer(num=3)
        self.answer_4_btn.bind('<Return>', lambda: self.choose_answer(num=3))

        # 次へボタン
        self.next_btn = ttk.Button(self)
        self.next_btn['text'] = '次へ(q)'
        self.next_btn['style'] = "default.TButton"
        self.next_btn['width'] = config.default_button_width
        self.next_btn['padding'] = config.default_button_padding
        self.next_btn['command'] = self.create_quest
        self.next_btn.bind('<Return>', self.create_quest)

        # 問題ラベル
        quest_label = tk.Label(self)
        # quest_label['width'] = 1000
        quest_label['font'] = config.message_font
        quest_label['text'] = "Q." 
        quest_label.place(x=20,y=40)

        # 問題メッセージ
        self.quest_message = tk.Message(self)
        self.quest_message['width'] = 980
        self.quest_message['font'] = config.message_font
        self.quest_message['text'] = "" 
        self.quest_message.place(x=40,y=40)

        # 解答ラベル1
        self.answer_label_1 = tk.Label(self)
        self.answer_label_1['font'] = config.message_font
        self.answer_label_1['text'] = "1." 
        self.answer_label_1.place(x=20,y=180)

        # 解答メッセージ1
        self.answer_message_1 = tk.Message(self)
        self.answer_message_1['width'] = 980
        self.answer_message_1['font'] = config.message_font
        self.answer_message_1['text'] = "" 
        self.answer_message_1.place(x=40,y=180)

        # 解答ラベル2
        self.answer_label_2 = tk.Label(self)
        self.answer_label_2['font'] = config.message_font
        self.answer_label_2['text'] = "2." 
        self.answer_label_2.place(x=20,y=300)

        # 解答メッセージ2
        self.answer_message_2 = tk.Message(self)
        self.answer_message_2['width'] = 980
        self.answer_message_2['font'] = config.message_font
        self.answer_message_2['text'] = "" 
        self.answer_message_2.place(x=40,y=300)

        # 解答ラベル3
        self.answer_label_3 = tk.Label(self)
        self.answer_label_3['font'] = config.message_font
        self.answer_label_3['text'] = "3." 
        self.answer_label_3.place(x=20,y=420)

        # 解答メッセージ3
        self.answer_message_3 = tk.Message(self)
        self.answer_message_3['width'] = 980
        self.answer_message_3['font'] = config.message_font
        self.answer_message_3['text'] = "" 
        self.answer_message_3.place(x=40,y=420)

        # 解答ラベル4
        self.answer_label_4 = tk.Label(self)
        self.answer_label_4['font'] = config.message_font
        self.answer_label_4['text'] = "4." 
        self.answer_label_4.place(x=20,y=540)

        # 解答メッセージ4
        self.answer_message_4 = tk.Message(self)
        self.answer_message_4['width'] = 980
        self.answer_message_4['font'] = config.message_font
        self.answer_message_4.place(x=40,y=540)

    def change_index(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_index(self.root)
        app.pack()
    
    # チェックボックスより章を決定
    def decision_chapter_number(self):
        while True:
            # 1～11の数字をランダムに生成
            rand_num = random.randint(1, 11)
            # self.is_checklistのrand_num番目がTrueの場合、rand_numを返す
            if self.is_checklist[rand_num - 1].get():
                print("選ばれた章は" + str(rand_num))
                return rand_num
    def yougo_to_imi_create(self):
        self.answer_message_1['text'] = config.ws[f'D{self.decision_answers[0]}'].value
        self.answer_message_2['text'] = config.ws[f'D{self.decision_answers[1]}'].value
        self.answer_message_3['text'] = config.ws[f'D{self.decision_answers[2]}'].value
        self.answer_message_4['text'] = config.ws[f'D{self.decision_answers[3]}'].value

        # 問題を一つだけ作成
        self.decision_quest = random.sample(self.decision_answers,1)
        self.quest_message['text'] = config.ws[f'B{self.decision_quest[0]}'].value + "　を説明したものはどれか？"
        print("今回の答えは" + str((self.decision_answers.index(self.decision_quest[0])) + 1))
    def imi_to_yougo_create(self):
        self.answer_message_1['text'] = config.ws[f'B{self.decision_answers[0]}'].value
        self.answer_message_2['text'] = config.ws[f'B{self.decision_answers[1]}'].value
        self.answer_message_3['text'] = config.ws[f'B{self.decision_answers[2]}'].value
        self.answer_message_4['text'] = config.ws[f'B{self.decision_answers[3]}'].value

        # 問題を一つだけ作成
        self.decision_quest = random.sample(self.decision_answers,1)
        self.quest_message['text'] = config.ws[f'D{self.decision_quest[0]}'].value + "　はどれを説明したものか？"
        print("今回の答えは" + str((self.decision_answers.index(self.decision_quest[0])) + 1))
    #問題を作成 
    def create_quest(self,event=None):
        # ボタン配置変更
        self.answer_1_btn.place(x='1058', y='40')
        self.answer_2_btn.place(x='1058', y='120')
        self.answer_3_btn.place(x='1058', y='200')
        self.answer_4_btn.place(x='1058', y='280')
        self.next_btn.place_forget()

        # 選択肢の色を戻す
        self.answer_label_1["fg"] = "black"
        self.answer_message_1["fg"] = "black"
        self.answer_label_2["fg"] = "black"
        self.answer_message_2["fg"] = "black"
        self.answer_label_3["fg"] = "black"
        self.answer_message_3["fg"] = "black"
        self.answer_label_4["fg"] = "black"
        self.answer_message_4["fg"] = "black"

        # ラベルを初期化
        self.label['text'] =""
        # 
        chpter_number = self.decision_chapter_number()
        # G列が選ばれた章の行番号を取得
        rows = []
        for i in range(1, config.ws.max_row + 1):
            if config.ws.cell(row=i, column=7).value == chpter_number:
                rows.append(i)

        # 一意に4つ作成
        self.decision_answers = random.sample(rows,4)

        # クエストモードにより取得列を変える
        if self.quest_mode == 1:
            self.yougo_to_imi_create()
        elif self.quest_mode ==2:
            self.imi_to_yougo_create()
        else:
            if random.randint(1,2) == 1:
                self.yougo_to_imi_create()
            else:
                self.imi_to_yougo_create()

    # 答えをクリック
    def choose_answer(self,num=None):
        # ボタン配置変更
        self.answer_1_btn.place_forget()
        self.answer_3_btn.place_forget()
        self.answer_2_btn.place_forget()
        self.answer_4_btn.place_forget()
        self.next_btn.place(x='1058', y='699')

        # 正解の選択肢の色を赤に
        if self.decision_answers.index(self.decision_quest[0]) == 0:
            self.answer_label_1["fg"] = "red"
            self.answer_message_1["fg"] = "red"
        elif self.decision_answers.index(self.decision_quest[0]) == 1:
            self.answer_label_2["fg"] = "red"
            self.answer_message_2["fg"] = "red"
        elif self.decision_answers.index(self.decision_quest[0]) == 2:
            self.answer_label_3["fg"] = "red"
            self.answer_message_3["fg"] = "red"
        else :
            self.answer_label_4["fg"] = "red"
            self.answer_message_4["fg"] = "red"
        
        # 正解判定
        if self.decision_answers[num] == self.decision_quest[0]:
            print("正解！")
            self.label['text'] = "正解！"
            self.label['fg'] ="red"
            quest_cell = config.ws['{}{}'.format('E',self.decision_quest[0])]
            quest_cell.value = quest_cell.value + 1
            config.wb.save(config.file_name)

        else:
            print("不正解・・・")
            self.label['text'] ="不正解・・・"
            self.label['fg'] ="blue"
            quest_cell = config.ws['{}{}'.format('F',self.decision_quest[0])]
            quest_cell.value = quest_cell.value + 1
            config.wb.save(config.file_name)

class Application_grades(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=1344, height=756, borderwidth=1, relief='solid')
        self.root = root
        self.pack(anchor='center',expand=1)
        self.pack_propagate(0)
        self.create_widgets()
        root.bind("m", lambda event: self.change_index() if self.return_btn.winfo_ismapped() else None)

        for i in range(11):
            self.calc_grades(i)

    def create_widgets(self):     
        style = ttk.Style()
        style.configure("default.TButton", font=("メイリオ", 16))
        
        # メニューに戻るボタン
        self.return_btn = ttk.Button(self)
        self.return_btn['text'] = 'メニューに戻る(m)'
        self.return_btn['style'] = "default.TButton"
        self.return_btn['width'] = config.default_button_width
        self.return_btn['padding'] = config.default_button_padding
        self.return_btn['command'] = self.change_index
        self.return_btn.bind('<Return>', self.change_index)
        self.return_btn.pack(side='bottom',anchor='sw') 

        # 線を引くための定義
        canvas = tk.Canvas(self, width=1500, height=800)
        canvas.pack()
        
        # トピックラベル
        chapter_label_0 = tk.Label(self)
        chapter_label_0['width'] = 32
        # chapter_label_0['relief'] = 'solid'
        chapter_label_0['text'] = "成績：チャプター"
        chapter_label_0['font'] = config.checkbtn_font
        chapter_label_0['anchor'] = tk.W
        chapter_label_0.place(x='40', y='20')

        chapter_label_0 = tk.Label(self)
        chapter_label_0['width'] = 10
        # chapter_label_0['relief'] = 'solid'
        chapter_label_0['text'] = "回答数"
        chapter_label_0['font'] = config.checkbtn_font
        chapter_label_0['anchor'] = tk.E 
        chapter_label_0.place(x='600', y='20')

        chapter_label_0 = tk.Label(self)
        chapter_label_0['width'] = 10
        # chapter_label_0['relief'] = 'solid'
        chapter_label_0['text'] = "正解率"
        chapter_label_0['font'] = config.checkbtn_font
        chapter_label_0['anchor'] = tk.E 
        chapter_label_0.place(x='800', y='20')

        chapter_label_0 = tk.Label(self)
        chapter_label_0['width'] = 20
        # chapter_label_0['relief'] = 'solid'
        chapter_label_0['text'] = "リセットボタン(未実装)"
        chapter_label_0['font'] = config.checkbtn_font
        chapter_label_0['anchor'] = tk.W
        chapter_label_0.place(x='1054', y='20')

        # 成績のウィジェット達
        self.number_of_answers_messages=[]
        self.accuracy_rate_messages=[]
        self.reset_btns = []

        for i in range(11):
            chapter_label = tk.Label(self)
            chapter_label['width'] = 32
            # chapter_label['relief'] = 'solid'
            chapter_label['text'] = config.chapter_list[i]
            chapter_label['font'] = config.checkbtn_font
            chapter_label['anchor'] = tk.W
            chapter_label.place(x='40', y=str(80 + i*55))

            number_of_answers_message = tk.Label(self)
            number_of_answers_message['width'] = 10
            # number_of_answers_message['relief'] = 'solid'
            number_of_answers_message['font'] = config.checkbtn_font
            number_of_answers_message['anchor'] = tk.E 
            number_of_answers_message.place(x='600', y=str(80 + i*55))

            accuracy_rate_message = tk.Label(self)
            accuracy_rate_message['width'] = 10
            # accuracy_rate_message['relief'] = 'solid'
            accuracy_rate_message['font'] = config.checkbtn_font
            accuracy_rate_message['anchor'] = tk.E 
            accuracy_rate_message.place(x='800', y=str(80 + i*55))

            reset_btn = ttk.Button(self)
            reset_btn['text'] = '第'+ str(i+1) +'章をリセットする'
            reset_btn['style'] = "default.TButton"
            reset_btn['width'] = config.default_button_width
            reset_btn['padding'] = config.default_button_padding
            # reset_btn['command'] = lambda: self.reset_btn_click(i)
            # reset_btn.bind('<Return>', self.reset_btn_click(i))
            reset_btn.place(x='1054',y=str(76 + i*55))

            self.number_of_answers_messages.append(number_of_answers_message)
            self.accuracy_rate_messages.append(accuracy_rate_message)
            self.reset_btns.append(reset_btn)
            
            canvas.create_line(0, 76 + i*55 , 1500, 76 + i*55 )

    def reset_btn_click(self,i) :
        num = str(i)
        print(num + "が押されました。")

    def calc_grades(self,i):
        grade_chapter_namber = i +1
        rows = []
        total_e = 0
        total_f = 0
        for j in range(1, config.ws.max_row + 1):
            if config.ws.cell(row=j, column=7).value == grade_chapter_namber:
                rows.append(j)

        for row in rows:
            total_e += config.ws.cell(row=row, column=5).value
            total_f += config.ws.cell(row=row, column=6).value

            number_of_answers = total_e + total_f
            accuracy_rate = total_e / (total_e + total_f) * 100 if number_of_answers > 0 else '0'
            self.number_of_answers_messages[i]["text"] = str(number_of_answers) + "回"
            self.accuracy_rate_messages[i]["text"] = str(int(accuracy_rate)) + "%"            

    def change_index(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_index(self.root)
        app.pack()

class Application_data(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=1344, height=756, borderwidth=1, relief='solid')
        self.root = root
        self.pack(anchor='center',expand=1)
        self.pack_propagate(0)
        self.create_widgets()
        root.bind("m", lambda event: self.change_index() if self.return_btn.winfo_ismapped() else None)

    def create_widgets(self):     
        style = ttk.Style()
        style.configure("default.TButton", font=("メイリオ", 16))
        
        # タイトルラベル
        label = tk.Label(self)
        label['text'] = 'ここはデータ変更ページ'
        label['font'] = config.titlefont
        label.pack()

        # メニューに戻るボタン
        self.return_btn = ttk.Button(self)
        self.return_btn['text'] = 'メニューに戻る(m)'
        self.return_btn['style'] = "default.TButton"
        self.return_btn['width'] = config.default_button_width
        self.return_btn['padding'] = config.default_button_padding
        self.return_btn['command'] = self.change_index
        self.return_btn.bind('<Return>', self.change_index)
        self.return_btn.pack(side='bottom',anchor='sw')        

    def change_index(self,event=None):
        # 現在のフレームを非表示にする
        self.pack_forget()
        # 新しいフレームを作成して表示する
        app = Application_index(self.root)
        app.pack()

root = tk.Tk()
root.title('FE単語意味頑張るマン')
root.geometry('1344x756')
# root.option_add("*size", 160)
# root.option_add("*foreground", "white")
# root.option_add("*background","black")
app = Application_index(root=root)
app.pack()
root.mainloop()
